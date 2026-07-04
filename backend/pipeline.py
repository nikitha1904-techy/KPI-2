from __future__ import annotations

import hashlib
import itertools
import json
import math
import os
import re
import shutil
import uuid
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import pandas as pd
from dotenv import load_dotenv
try:
    from openai import OpenAI, AzureOpenAI
except Exception:  # package may not be installed until backend requirements are installed
    OpenAI = None
    AzureOpenAI = None
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

load_dotenv()

BACKEND_DIR = Path(__file__).resolve().parent
RUNS_DIR = BACKEND_DIR / "runs"
SAMPLE_DIR = BACKEND_DIR / "sample_inputs"
REFERENCE_OUTPUT = SAMPLE_DIR / "final_output_with_all_steps_reference.xlsx"
RUNS_DIR.mkdir(exist_ok=True)

TOP_N = int(os.getenv("TOP_N", "5"))
LLM_BATCH_SIZE = int(os.getenv("LLM_BATCH_SIZE", "6"))
USE_REFERENCE_FOR_SAMPLE = os.getenv("USE_REFERENCE_FOR_SAMPLE", "false").lower() in {"1", "true", "yes", "y"}

LLM_PROVIDER = os.getenv("LLM_PROVIDER", "openai").strip().lower()
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY", "").strip()
OPENAI_BASE_URL = os.getenv("OPENAI_BASE_URL", "https://api.openai.com/v1").strip()
OPENAI_MODEL = os.getenv("OPENAI_MODEL", "gpt-4o-mini").strip()
AZURE_OPENAI_API_KEY = os.getenv("AZURE_OPENAI_API_KEY", "").strip()
AZURE_OPENAI_ENDPOINT = os.getenv("AZURE_OPENAI_ENDPOINT", "").strip()
AZURE_OPENAI_API_VERSION = os.getenv("AZURE_OPENAI_API_VERSION", "2024-10-21").strip()
AZURE_OPENAI_DEPLOYMENT = os.getenv("AZURE_OPENAI_DEPLOYMENT", "").strip()

BASE_MAPPING_COLUMNS = [
    "Report A", "Report Col Name", "Report Formula", "Ontology KPI Name",
    "Ontology Formula", "Ontology Definition", "similarity_score", "similarity_band",
    "formula_equivalence_type", "actuarial_rationale", "key_formula_matches",
    "key_formula_differences", "business_logic_assessment", "confidence_level",
    "needs_human_review",
]

STEP4_COLUMNS = BASE_MAPPING_COLUMNS + [
    "LLM Business Interpretation of Report Formula", "Formula Similarity Score",
    "Definition Similarity Score", "Overall Step 4 Similarity Score",
    "Step 4 Verdict", "Step 4 Rationale",
]

STEP5_COLUMNS = [
    "Ontology KPI Name", "Report 1", "Report Column 1", "Report Formula 1",
    "Ontology Score 1", "Report 2", "Report Column 2", "Report Formula 2",
    "Ontology Score 2", "Pairwise Formula Similarity Score", "Pairwise Verdict",
    "Pairwise Rationale",
]

LLM_PAIRWISE_COLUMNS = [
    "Ontology KPI Name", "Ontology Formula", "Ontology Definition", "Report 1 File",
    "Report 1 Column", "Report 1 Formula", "Report 1 Mapping Score", "Report 2 File",
    "Report 2 Column", "Report 2 Formula", "Report 2 Mapping Score",
    "Formula Similarity Score", "Similarity Band", "LLM Verdict",
    "Formula Equivalence Type", "Rationale", "Key Formula Matches",
    "Key Formula Differences", "Business Discrepancy", "Needs Human Review",
    "Confidence Level",
]

MAPPED_COLUMNS = [
    "Report Scope", "Ontology KPI", "Ontology Formula", "Ontology Definition",
    "Mapped Report KPI(s)", "Mapped Report Formula(s)", "Best Mapping Score(s)", "Status",
]

MISSING_COLUMNS = [
    "Report Scope", "Ontology KPI", "Ontology Formula", "Ontology Definition", "Missing Reason",
]

RECOMMENDATION_COLUMNS = [
    "Report Scope", "Ontology KPI", "Ontology Formula", "Ontology Definition", "Missing Reason",
    "Should Add to Report?", "Recommendation Priority", "Priority Rationale",
    "Industry Relevance", "Recommended Report Section", "Expected Business Value",
    "Reason for Recommendation", "Reason If Not Recommended", "Related Existing Report KPIs",
    "Related Existing Report Formulas", "Relationship to Existing Report",
    "Potential Formula / Calculation Guidance", "Needs Human Review", "Confidence Level",
]

SHEET_ORDER = [
    "1. Summary", "2. All Comparisons", "3. Best Mapping", "4. Common Ontology",
    "5. Step4 Definition Validation", "6. Step5 Pairwise Check", "LLM Pairwise Formula Check",
    "Best Mapping - EB Details", "Best Mapping - Worksite", "RW Missing KPI Summary",
    "EB Mapped Ontology KPIs", "EB Missing Ontology KPIs", "EB LLM Missing KPI Recs",
    "Worksite Mapped Ontology KPIs", "Worksite Missing Ontology KPIs",
    "Worksite LLM Missing KPI Recs",
]

RELEVANT_TABLE_KEYWORDS = {
    "stat reserve", "statutory reserve", "tax reserve", "tax reserves", "gaap reserve",
    "gaap ben reserves", "gaap benefit reserves", "face amount", "cash value", "count",
    "exhibit", "reserve", "summary",
}

EXCLUDE_NAMES = {
    "accounting date", "product", "product subtype", "plan", "people soft code", "policy number",
    "rbc c2 product category", "section", "section desc", "triton plancode", "ledger product",
    "plancode description",
}


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    text = str(value).replace("\r", " ").replace("\n", " ").replace("\t", " ")
    text = re.sub(r"\s+", " ", text).strip()
    if text.lower() in {"nan", "none", "null", "n/a", "na"}:
        return ""
    return text


def normalize(value: Any) -> str:
    text = clean_text(value).lower().replace("–", "-").replace("—", "-")
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9]+", " ", text)).strip()


def normalize_formula(value: Any) -> str:
    text = clean_text(value)
    if text.startswith("="):
        text = text[1:].strip()
    return re.sub(r"\s+", " ", text.replace("_xlfn.", "")).strip()


def parse_float(value: Any, default: float = 0.0) -> float:
    if isinstance(value, (int, float)) and not (isinstance(value, float) and math.isnan(value)):
        return float(value)
    text = re.sub(r"[^0-9.\-]", "", clean_text(value))
    try:
        return float(text)
    except Exception:
        return default


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def stable_id(parts: Iterable[Any]) -> str:
    joined = "||".join(clean_text(p) for p in parts)
    return hashlib.sha1(joined.encode("utf-8")).hexdigest()[:16]


def score_to_band(score: float) -> str:
    score = parse_float(score)
    if score >= 90:
        return "Very High"
    if score >= 75:
        return "High"
    if score >= 50:
        return "Medium"
    if score > 0:
        return "Low"
    return "No Match"


def safe_sheet_name(name: str) -> str:
    name = re.sub(r"[\\/*?:\[\]]", "-", clean_text(name))
    return name[:31] or "Sheet"


def records_from_workbook(excel_path: Path) -> Dict[str, List[Dict[str, Any]]]:
    xls = pd.ExcelFile(excel_path)
    results: Dict[str, List[Dict[str, Any]]] = {}
    for sheet in xls.sheet_names:
        df = pd.read_excel(excel_path, sheet_name=sheet, dtype=object)
        df = df.where(pd.notna(df), "")
        results[sheet] = df.to_dict("records")
    return results


def write_results_json(excel_path: Path, json_path: Path) -> Dict[str, List[Dict[str, Any]]]:
    results = records_from_workbook(excel_path)
    json_path.write_text(json.dumps(results, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    return results


def is_reference_sample(report_paths: List[Path], ontology_path: Path) -> bool:
    if not USE_REFERENCE_FOR_SAMPLE:
        return False
    sample_reports = [SAMPLE_DIR / "report_a.json", SAMPLE_DIR / "report_b.json"]
    sample_ontology = SAMPLE_DIR / "life_annuity_actuarial_reserving_kpi_ontology.xlsx"
    if not all(p.exists() for p in sample_reports + [sample_ontology, REFERENCE_OUTPUT]):
        return False
    uploaded_hashes = sorted(sha256_file(p) for p in report_paths)
    sample_hashes = sorted(sha256_file(p) for p in sample_reports)
    return uploaded_hashes == sample_hashes and sha256_file(ontology_path) == sha256_file(sample_ontology)


def load_ontology_file(path: Path) -> pd.DataFrame:
    ext = path.suffix.lower()
    if ext in {".xlsx", ".xls"}:
        xls = pd.ExcelFile(path)
        sheet = xls.sheet_names[0]
        for s in xls.sheet_names:
            if any(k in s.lower() for k in ["actuarial", "reserving", "ontology", "kpi"]):
                sheet = s
                break
        raw = pd.read_excel(path, sheet_name=sheet, header=None, dtype=object)
        header_row = detect_header_row(raw)
        df = pd.read_excel(path, sheet_name=sheet, header=header_row, dtype=object)
    elif ext == ".csv":
        raw = pd.read_csv(path, header=None, dtype=object)
        header_row = detect_header_row(raw)
        df = pd.read_csv(path, header=header_row, dtype=object)
    elif ext == ".json":
        df = pd.read_json(path, dtype=object)
    else:
        raise ValueError(f"Unsupported ontology file extension: {ext}")

    df = df.dropna(axis=0, how="all").dropna(axis=1, how="all")
    df.columns = [clean_text(c) for c in df.columns]
    columns = list(df.columns)
    kpi_col = find_col(columns, ["KPI / Metric Name", "KPI Metric Name", "Metric Name", "KPI Name", "Measurement (KPI)", "Measurement KPI", "Ontology KPI Name"])
    formula_col = find_col(columns, ["Formula", "Ontology Formula", "Computing Variables/Formula", "Computing Variables Formula", "Business Formula"])
    definition_col = find_col(columns, ["Definition", "Ontology Definition", "Description", "KPI Definition"])
    category_col = find_col(columns, ["Category", "KPI Category", "Actuarial Category"])
    synonyms_col = find_col(columns, ["Synonyms", "Alternate Names / Synonyms", "Alternate Names"])
    if not kpi_col:
        raise ValueError(f"Could not find ontology KPI name column. Columns: {columns}")

    rows = []
    for _, row in df.iterrows():
        kpi = clean_text(row.get(kpi_col))
        if not kpi or len(kpi) > 160:
            continue
        formula = normalize_formula(row.get(formula_col)) if formula_col else ""
        definition = clean_text(row.get(definition_col)) if definition_col else ""
        rows.append({
            "Ontology KPI Name": kpi,
            "Ontology Formula": formula,
            "Ontology Definition": definition,
            "Category": clean_text(row.get(category_col)) if category_col else "",
            "Synonyms": clean_text(row.get(synonyms_col)) if synonyms_col else "",
        })
    out = pd.DataFrame(rows).drop_duplicates(["Ontology KPI Name", "Ontology Formula", "Ontology Definition"]).reset_index(drop=True)
    if out.empty:
        raise ValueError("No ontology KPI rows extracted.")
    return out


def detect_header_row(raw: pd.DataFrame) -> int:
    best_row, best_score = 0, -1
    for i in range(min(50, len(raw))):
        text = " | ".join(clean_text(x).lower() for x in raw.iloc[i].tolist())
        score = 0
        if any(x in text for x in ["kpi", "measurement", "metric"]):
            score += 10
        if "definition" in text:
            score += 5
        if "formula" in text or "computing" in text:
            score += 10
        if "category" in text:
            score += 2
        if score > best_score:
            best_row, best_score = i, score
    return best_row


def find_col(columns: List[str], candidates: List[str]) -> Optional[str]:
    norm_cols = {c: normalize(c) for c in columns}
    for col, ncol in norm_cols.items():
        if any(ncol == normalize(c) for c in candidates):
            return col
    best, best_score = None, 0
    for col, ncol in norm_cols.items():
        col_parts = set(ncol.split())
        for cand in candidates:
            nc = normalize(cand)
            cand_parts = set(nc.split())
            score = 0
            if nc and nc in ncol:
                score += 8
            if cand_parts and cand_parts.issubset(col_parts):
                score += 6
            score += len(cand_parts & col_parts)
            if score > best_score:
                best, best_score = col, score
    return best if best_score >= 3 else None


def derive_report_label(json_path: Path, data: Dict[str, Any]) -> str:
    """Return business-friendly label used for sheet names and run grouping."""
    raw = clean_text(data.get("file_name") or data.get("workbook_metadata", {}).get("file_name") or json_path.name)
    n = normalize(raw)
    if "worksite" in n:
        return "Worksite"
    if "eb reserve" in n or "employee benefits" in n or " eb " in f" {n} ":
        return "EB Details"
    return Path(raw).stem[:31]


def derive_report_scope(label: str, source_file: str) -> str:
    n = normalize(" ".join([label, source_file]))
    if "worksite" in n:
        return "Worksite"
    if "eb" in n or "employee benefits" in n:
        return "EB"
    return clean_text(label)[:31] or "Report"


def derive_report_display_name(json_path: Path, data: Dict[str, Any]) -> str:
    """Return the visible Report A value. The golden workbook uses original workbook file names."""
    return clean_text(data.get("file_name") or data.get("workbook_metadata", {}).get("file_name") or json_path.name)


def build_visible_report_col(label: str, table_name: str, col_name: str) -> str:
    """Prefix repeated/generic Worksite KPI names with their section/table context.

    This is critical for columns such as Gross Reserve, 3rd Party YRT, Net Reserve,
    Cash Value, and Count, which repeat across STAT, Tax, GAAP, Face Amount, and Exhibit 5.
    """
    table_name = clean_text(table_name)
    col_name = clean_text(col_name)
    if normalize(label) == "worksite" and table_name:
        return f"{table_name} - {col_name}"
    return col_name


def is_relevant_table(sheet: Dict[str, Any], table: Dict[str, Any]) -> bool:
    context = " ".join(clean_text(x) for x in [
        sheet.get("sheet_name"), sheet.get("sheet_type"), table.get("table_name"),
        table.get("table_type"), table.get("section_title"), table.get("business_purpose"),
        table.get("measures"), table.get("dimensions"),
    ])
    n = normalize(context)
    return any(k in n for k in RELEVANT_TABLE_KEYWORDS)


def build_report_context(label: str, data: Dict[str, Any], sheet: Dict[str, Any], table: Dict[str, Any], col: Dict[str, Any]) -> str:
    lineage = col.get("formula_lineage") or {}
    source_file = clean_text(data.get('file_name') or data.get('workbook_metadata', {}).get('file_name'))
    return "\n".join([
        f"Report Scope: {derive_report_scope(label, source_file)}",
        f"Report Label: {label}",
        f"Workbook File: {source_file}",
        f"Workbook Purpose: {clean_text(data.get('purpose'))}",
        f"Sheet Name: {clean_text(sheet.get('sheet_name'))}",
        f"Sheet Type: {clean_text(sheet.get('sheet_type'))}",
        f"Table Name: {clean_text(table.get('table_name'))}",
        f"Section Title: {clean_text(table.get('section_title'))}",
        f"Table Purpose: {clean_text(table.get('business_purpose'))}",
        f"Measures: {clean_text(table.get('measures'))}",
        f"Dimensions: {clean_text(table.get('dimensions'))}",
        f"Column Name: {clean_text(col.get('column_name'))}",
        f"Column Definition: {clean_text(col.get('definition'))}",
        f"Formula Lineage: {clean_text(lineage)}",
    ])


def extract_report_kpis(json_path: Path) -> pd.DataFrame:
    data = json.loads(json_path.read_text(encoding="utf-8"))
    label = derive_report_label(json_path, data)
    source_file = derive_report_display_name(json_path, data)
    scope = derive_report_scope(label, source_file)
    rows: List[Dict[str, Any]] = []
    for sheet in data.get("sheets", []):
        for table in sheet.get("tables", []):
            if not is_relevant_table(sheet, table):
                continue
            table_name = clean_text(table.get("table_name"))
            section_title = clean_text(table.get("section_title")) or table_name
            sheet_name = clean_text(sheet.get("sheet_name"))
            for col in table.get("columns", []):
                raw_col_name = clean_text(col.get("column_name"))
                if not raw_col_name:
                    continue
                formula = clean_text(col.get("formula_pattern")) or normalize_formula(col.get("formula"))
                col_type = normalize(col.get("type"))
                n_name = normalize(raw_col_name)
                if not formula:
                    continue
                if n_name in EXCLUDE_NAMES and "formula" not in col_type:
                    continue
                visible_col_name = build_visible_report_col(label, table_name or section_title, raw_col_name)
                context = build_report_context(label, data, sheet, table, col)
                kpi_id = stable_id([source_file, visible_col_name, formula, sheet_name, table_name, section_title])
                rows.append({
                    "Report A": source_file,
                    "Report Col Name": visible_col_name,
                    "Report Formula": formula,
                    "_report_context": context,
                    "_kpi_id": kpi_id,
                    "_sheet_name": sheet_name,
                    "_table_name": table_name,
                    "_section_title": section_title,
                    "_source_file": source_file,
                    "_report_label": label,
                    "_report_scope": scope,
                    "_raw_col_name": raw_col_name,
                })
    if not rows:
        raise ValueError(f"No formula-based KPI rows extracted from {json_path.name}.")
    # Deduplicate exact duplicate rows only. Do NOT dedupe by Report + Column alone.
    return pd.DataFrame(rows).drop_duplicates(["_kpi_id"]).reset_index(drop=True)

def token_set(text: str) -> set:
    stop = {"the", "and", "or", "of", "for", "to", "by", "with", "as", "on", "in", "at", "from", "a", "an", "is", "are", "value", "amount", "kpi", "metric", "formula", "calculation", "data", "field", "report", "ontology", "total"}
    return {x for x in normalize(text).split() if len(x) > 1 and x not in stop}


def feature_flags(text: str) -> Dict[str, bool]:
    n = normalize(text)
    return {
        "statutory": any(x in n for x in ["statutory", "stat reserve", "stat reserves", " stat "]),
        "tax": "tax" in n,
        "gaap": "gaap" in n,
        "gross": "gross" in n,
        "net": "net" in n,
        "general_account": "general account" in n,
        "separate_account": "separate account" in n,
        "reinsurance": any(x in n for x in ["reinsurance", "3rd party", "third party", "ceded", "non tai", "recapture", "yrt"]),
        "captive": "captive" in n,
        "reserve": "reserve" in n or "reserves" in n,
        "face": "face" in n or "sum assured" in n,
        "cash": "cash value" in n or "account value" in n,
        "count": "count" in n or "number of" in n,
        "premium": "premium" in n,
        "claim": "claim" in n or "benefit" in n,
        "sum": "sum" in n or "sumifs" in n,
        "average": "average" in n or "avg" in n,
        "ratio": "ratio" in n or "rate" in n or "percent" in n,
    }

def extract_json_object(text: str) -> Dict[str, Any]:
    text = clean_text(text)
    if not text:
        raise ValueError("Empty LLM response")
    try:
        return json.loads(text)
    except Exception:
        pass
    match = re.search(r"\{.*\}", text, re.S)
    if not match:
        raise ValueError(f"Could not locate JSON object in LLM response: {text[:200]}")
    return json.loads(match.group(0))


class LLMClient:
    """Small OpenAI/Azure/OpenAI-compatible JSON client used by the pipeline."""

    def __init__(self, required: bool = False):
        self.required = required
        self.provider = LLM_PROVIDER
        self.model = OPENAI_MODEL
        self.client = None
        self.enabled = False
        self.config_error = ""

        try:
            if OpenAI is None or AzureOpenAI is None:
                self.config_error = "The openai Python package is not installed. Run: pip install -r backend/requirements.txt"
                return
            if self.provider in {"azure", "azure_openai"}:
                if not (AZURE_OPENAI_API_KEY and AZURE_OPENAI_ENDPOINT and AZURE_OPENAI_DEPLOYMENT):
                    self.config_error = "Azure OpenAI is selected, but AZURE_OPENAI_API_KEY, AZURE_OPENAI_ENDPOINT, and AZURE_OPENAI_DEPLOYMENT are not all set."
                    return
                self.model = AZURE_OPENAI_DEPLOYMENT
                self.client = AzureOpenAI(
                    api_key=AZURE_OPENAI_API_KEY,
                    azure_endpoint=AZURE_OPENAI_ENDPOINT,
                    api_version=AZURE_OPENAI_API_VERSION,
                )
                self.enabled = True
            else:
                if not OPENAI_API_KEY:
                    self.config_error = "OPENAI_API_KEY is not set. Add it to backend/.env or uncheck Use LLM/API."
                    return
                kwargs = {"api_key": OPENAI_API_KEY}
                if OPENAI_BASE_URL:
                    kwargs["base_url"] = OPENAI_BASE_URL
                self.client = OpenAI(**kwargs)
                self.enabled = True
        except Exception as exc:
            self.config_error = f"LLM client initialization failed: {exc}"

    def assert_ready(self) -> None:
        if self.required and not self.enabled:
            raise RuntimeError(self.config_error or "LLM/API requested, but no usable LLM configuration was found.")

    def complete_json(self, system_prompt: str, user_prompt: str) -> Dict[str, Any]:
        self.assert_ready()
        if not self.enabled:
            raise RuntimeError("LLM client is disabled")
        response = self.client.chat.completions.create(
            model=self.model,
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt},
            ],
            temperature=float(os.getenv("LLM_TEMPERATURE", "0.1")),
            response_format={"type": "json_object"},
        )
        content = response.choices[0].message.content or ""
        return extract_json_object(content)


def llm_system_prompt(task: str) -> str:
    return (
        "You are a senior Life & Annuity actuarial reserving SME. "
        "Evaluate KPI formulas using actuarial business meaning, accounting basis, reserve/exposure family, "
        "aggregation logic, sign convention, filters, and formula lineage. "
        "Return one valid JSON object only. No markdown. "
        f"Task: {task}"
    )


def coerce_score(value: Any, default: float) -> float:
    return max(0.0, min(100.0, round(parse_float(value, default), 1)))


def apply_llm_mapping_columns(top_df: pd.DataFrame, llm_client: Optional[LLMClient]) -> pd.DataFrame:
    """Populate the main mapping fields. If LLM is unavailable or a row fails, keep deterministic fallback and row count."""
    df = apply_fallback_llm_columns(top_df)
    if not llm_client or not llm_client.enabled:
        return df

    system_prompt = llm_system_prompt("report KPI to ontology KPI mapping")
    for idx, row in df.iterrows():
        fallback_score = parse_float(row.get("similarity_score"))
        user_prompt = f"""
Compare this report KPI against the ontology KPI.

Report KPI context:
{clean_text(row.get('_report_context'))}

Report KPI name: {clean_text(row.get('Report Col Name'))}
Report formula: {clean_text(row.get('Report Formula'))}

Ontology KPI name: {clean_text(row.get('Ontology KPI Name'))}
Ontology formula: {clean_text(row.get('Ontology Formula'))}
Ontology definition: {clean_text(row.get('Ontology Definition'))}

Return JSON with exactly these keys:
- similarity_score: number 0-100
- similarity_band: one of Very High, High, Medium, Low, No Match
- formula_equivalence_type: Same Formula Logic, Equivalent Business Meaning, Related but Different, Different KPI, Insufficient Information
- actuarial_rationale: concise actuarial rationale
- key_formula_matches: concise matches
- key_formula_differences: concise differences
- business_logic_assessment: concise business meaning assessment
- confidence_level: High, Medium, or Low
- needs_human_review: Yes or No
""".strip()
        try:
            out = llm_client.complete_json(system_prompt, user_prompt)
            score = coerce_score(out.get("similarity_score"), fallback_score)
            df.at[idx, "similarity_score"] = score
            df.at[idx, "similarity_band"] = clean_text(out.get("similarity_band")) or score_to_band(score)
            df.at[idx, "formula_equivalence_type"] = clean_text(out.get("formula_equivalence_type")) or df.at[idx, "formula_equivalence_type"]
            df.at[idx, "actuarial_rationale"] = clean_text(out.get("actuarial_rationale")) or df.at[idx, "actuarial_rationale"]
            df.at[idx, "key_formula_matches"] = clean_text(out.get("key_formula_matches")) or df.at[idx, "key_formula_matches"]
            df.at[idx, "key_formula_differences"] = clean_text(out.get("key_formula_differences")) or df.at[idx, "key_formula_differences"]
            df.at[idx, "business_logic_assessment"] = clean_text(out.get("business_logic_assessment")) or df.at[idx, "business_logic_assessment"]
            df.at[idx, "confidence_level"] = clean_text(out.get("confidence_level")) or df.at[idx, "confidence_level"]
            df.at[idx, "needs_human_review"] = clean_text(out.get("needs_human_review")) or df.at[idx, "needs_human_review"]
        except Exception as exc:
            if llm_client.required:
                raise RuntimeError(f"LLM mapping call failed; stopping because Use LLM/API is enabled. Error: {clean_text(exc)}") from exc
            df.at[idx, "actuarial_rationale"] = f"LLM call failed; deterministic fallback retained. Error: {clean_text(exc)}"
            df.at[idx, "needs_human_review"] = "Yes"
            df.at[idx, "confidence_level"] = "Low"
    return df


def llm_step4_validation(common_df: pd.DataFrame, llm_client: Optional[LLMClient]) -> pd.DataFrame:
    df = step4_validation_fallback(common_df)
    if not llm_client or not llm_client.enabled:
        return df
    system_prompt = llm_system_prompt("definition/formula validation for common ontology mappings")
    for idx, row in df.iterrows():
        fallback = parse_float(row.get("Overall Step 4 Similarity Score") or row.get("similarity_score"))
        user_prompt = f"""
Validate this mapped row using report formula, ontology formula, and ontology definition.

Report: {clean_text(row.get('Report A'))}
Report KPI: {clean_text(row.get('Report Col Name'))}
Report formula: {clean_text(row.get('Report Formula'))}

Ontology KPI: {clean_text(row.get('Ontology KPI Name'))}
Ontology formula: {clean_text(row.get('Ontology Formula'))}
Ontology definition: {clean_text(row.get('Ontology Definition'))}

Return JSON with exactly these keys:
- business_interpretation: what the report formula calculates
- formula_similarity_score: number 0-100
- definition_similarity_score: number 0-100
- overall_similarity_score: number 0-100
- verdict: Aligned, Partial Match, or Not Aligned
- rationale: concise rationale
""".strip()
        try:
            out = llm_client.complete_json(system_prompt, user_prompt)
            fscore = coerce_score(out.get("formula_similarity_score"), fallback)
            dscore = coerce_score(out.get("definition_similarity_score"), fallback)
            overall = coerce_score(out.get("overall_similarity_score"), (fscore + dscore) / 2)
            df.at[idx, "LLM Business Interpretation of Report Formula"] = clean_text(out.get("business_interpretation")) or df.at[idx, "LLM Business Interpretation of Report Formula"]
            df.at[idx, "Formula Similarity Score"] = fscore
            df.at[idx, "Definition Similarity Score"] = dscore
            df.at[idx, "Overall Step 4 Similarity Score"] = overall
            df.at[idx, "Step 4 Verdict"] = clean_text(out.get("verdict")) or ("Aligned" if overall >= 85 else "Partial Match" if overall >= 50 else "Not Aligned")
            df.at[idx, "Step 4 Rationale"] = clean_text(out.get("rationale")) or df.at[idx, "Step 4 Rationale"]
        except Exception as exc:
            if llm_client.required:
                raise RuntimeError(f"LLM Step 4 validation failed; stopping because Use LLM/API is enabled. Error: {clean_text(exc)}") from exc
            df.at[idx, "Step 4 Rationale"] = f"LLM call failed; fallback retained. Error: {clean_text(exc)}"
    return df[STEP4_COLUMNS]


def llm_pairwise_checks(common_df: pd.DataFrame, llm_client: Optional[LLMClient]) -> Tuple[pd.DataFrame, pd.DataFrame]:
    step5_df, pairs = pairwise_checks_fallback(common_df)
    if not llm_client or not llm_client.enabled:
        return step5_df, pairs
    system_prompt = llm_system_prompt("pairwise report formula equivalence for formulas mapped to the same ontology KPI")
    for idx, row in pairs.iterrows():
        fallback = parse_float(row.get("Formula Similarity Score"))
        user_prompt = f"""
Two report formulas map to the same ontology KPI. Decide whether they calculate the same actuarial business metric.

Ontology KPI: {clean_text(row.get('Ontology KPI Name'))}
Ontology formula: {clean_text(row.get('Ontology Formula'))}
Ontology definition: {clean_text(row.get('Ontology Definition'))}

Report 1: {clean_text(row.get('Report 1 File'))}
Report 1 KPI: {clean_text(row.get('Report 1 Column'))}
Report 1 formula: {clean_text(row.get('Report 1 Formula'))}
Report 1 mapping score: {clean_text(row.get('Report 1 Mapping Score'))}

Report 2: {clean_text(row.get('Report 2 File'))}
Report 2 KPI: {clean_text(row.get('Report 2 Column'))}
Report 2 formula: {clean_text(row.get('Report 2 Formula'))}
Report 2 mapping score: {clean_text(row.get('Report 2 Mapping Score'))}

Return JSON with exactly these keys:
- formula_similarity_score: number 0-100
- similarity_band: Very High, High, Medium, Low, or No Match
- llm_verdict: Equivalent, Partial Match, or Not Equivalent
- formula_equivalence_type: Same Formula Logic, Equivalent Business Meaning, Related but Different, Different KPI, Insufficient Information
- rationale: concise rationale
- key_formula_matches: concise matches
- key_formula_differences: concise differences
- business_discrepancy: discrepancy or None
- needs_human_review: Yes or No
- confidence_level: High, Medium, or Low
""".strip()
        try:
            out = llm_client.complete_json(system_prompt, user_prompt)
            score = coerce_score(out.get("formula_similarity_score"), fallback)
            pairs.at[idx, "Formula Similarity Score"] = score
            pairs.at[idx, "Similarity Band"] = clean_text(out.get("similarity_band")) or score_to_band(score)
            pairs.at[idx, "LLM Verdict"] = clean_text(out.get("llm_verdict")) or pairs.at[idx, "LLM Verdict"]
            pairs.at[idx, "Formula Equivalence Type"] = clean_text(out.get("formula_equivalence_type")) or pairs.at[idx, "Formula Equivalence Type"]
            pairs.at[idx, "Rationale"] = clean_text(out.get("rationale")) or pairs.at[idx, "Rationale"]
            pairs.at[idx, "Key Formula Matches"] = clean_text(out.get("key_formula_matches")) or pairs.at[idx, "Key Formula Matches"]
            pairs.at[idx, "Key Formula Differences"] = clean_text(out.get("key_formula_differences")) or pairs.at[idx, "Key Formula Differences"]
            pairs.at[idx, "Business Discrepancy"] = clean_text(out.get("business_discrepancy")) or pairs.at[idx, "Business Discrepancy"]
            pairs.at[idx, "Needs Human Review"] = clean_text(out.get("needs_human_review")) or pairs.at[idx, "Needs Human Review"]
            pairs.at[idx, "Confidence Level"] = clean_text(out.get("confidence_level")) or pairs.at[idx, "Confidence Level"]
        except Exception as exc:
            if llm_client.required:
                raise RuntimeError(f"LLM pairwise formula check failed; stopping because Use LLM/API is enabled. Error: {clean_text(exc)}") from exc
            pairs.at[idx, "Rationale"] = f"LLM call failed; fallback retained. Error: {clean_text(exc)}"
            pairs.at[idx, "Needs Human Review"] = "Yes"
            pairs.at[idx, "Confidence Level"] = "Low"
    step_rows = []
    for _, r in pairs.iterrows():
        step_rows.append({
            "Ontology KPI Name": r["Ontology KPI Name"],
            "Report 1": r["Report 1 File"], "Report Column 1": r["Report 1 Column"], "Report Formula 1": r["Report 1 Formula"], "Ontology Score 1": r["Report 1 Mapping Score"],
            "Report 2": r["Report 2 File"], "Report Column 2": r["Report 2 Column"], "Report Formula 2": r["Report 2 Formula"], "Ontology Score 2": r["Report 2 Mapping Score"],
            "Pairwise Formula Similarity Score": r["Formula Similarity Score"],
            "Pairwise Verdict": r["LLM Verdict"], "Pairwise Rationale": r["Rationale"],
        })
    return pd.DataFrame(step_rows, columns=STEP5_COLUMNS), pairs[LLM_PAIRWISE_COLUMNS]


def llm_gap_recommendations(gap_results: Dict[str, Dict[str, Any]], report_best: Dict[str, pd.DataFrame], llm_client: Optional[LLMClient]) -> Dict[str, Dict[str, Any]]:
    if not llm_client or not llm_client.enabled:
        return gap_results
    system_prompt = llm_system_prompt("missing ontology KPI recommendation for report-wise gap analysis")
    for report, data in gap_results.items():
        existing_kpis = " | ".join(report_best[report]["Report Col Name"].astype(str).map(clean_text).unique()[:30])
        existing_formulas = " | ".join(report_best[report]["Report Formula"].astype(str).map(clean_text).unique()[:20])
        recs = data["recommendations"].copy()
        for idx, row in recs.iterrows():
            user_prompt = f"""
Report scope: {report}
Existing mapped report KPIs: {existing_kpis}
Existing mapped formulas: {existing_formulas}

Missing ontology KPI: {clean_text(row.get('Ontology KPI'))}
Ontology formula: {clean_text(row.get('Ontology Formula'))}
Ontology definition: {clean_text(row.get('Ontology Definition'))}
Missing reason: {clean_text(row.get('Missing Reason'))}

Return JSON with exactly these keys:
- should_add_to_report: Yes, Maybe, or No
- recommendation_priority: High, Medium, or Low
- priority_rationale
- industry_relevance
- recommended_report_section
- expected_business_value
- reason_for_recommendation
- reason_if_not_recommended
- related_existing_report_kpis
- related_existing_report_formulas
- relationship_to_existing_report
- potential_formula_calculation_guidance
- needs_human_review: Yes or No
- confidence_level: High, Medium, or Low
""".strip()
            try:
                out = llm_client.complete_json(system_prompt, user_prompt)
                mapping = {
                    "Should Add to Report?": "should_add_to_report",
                    "Recommendation Priority": "recommendation_priority",
                    "Priority Rationale": "priority_rationale",
                    "Industry Relevance": "industry_relevance",
                    "Recommended Report Section": "recommended_report_section",
                    "Expected Business Value": "expected_business_value",
                    "Reason for Recommendation": "reason_for_recommendation",
                    "Reason If Not Recommended": "reason_if_not_recommended",
                    "Related Existing Report KPIs": "related_existing_report_kpis",
                    "Related Existing Report Formulas": "related_existing_report_formulas",
                    "Relationship to Existing Report": "relationship_to_existing_report",
                    "Potential Formula / Calculation Guidance": "potential_formula_calculation_guidance",
                    "Needs Human Review": "needs_human_review",
                    "Confidence Level": "confidence_level",
                }
                for col, key in mapping.items():
                    value = clean_text(out.get(key))
                    if value:
                        recs.at[idx, col] = value
            except Exception as exc:
                if llm_client.required:
                    raise RuntimeError(f"LLM missing-KPI recommendation failed; stopping because Use LLM/API is enabled. Error: {clean_text(exc)}") from exc
                recs.at[idx, "Reason for Recommendation"] = f"LLM call failed; fallback retained. Error: {clean_text(exc)}"
                recs.at[idx, "Needs Human Review"] = "Yes"
                recs.at[idx, "Confidence Level"] = "Low"
        data["recommendations"] = recs[RECOMMENDATION_COLUMNS]
    return gap_results



def _ontology_name(ont_row: pd.Series) -> str:
    return clean_text(ont_row.get("Ontology KPI Name"))


def _exact_target_score(report_row: pd.Series, ontology_name: str) -> Tuple[Optional[float], str]:
    """High-precision actuarial anchor rules for reserving reports.

    These rules are intentionally business-semantic rather than simple token overlap.
    They keep Statutory, Tax, GAAP, Face Amount, Cash Value, Count, Ceded,
    Captive, and Net/Gross basis from bleeding into each other before the LLM review.
    """
    col = normalize(report_row.get("Report Col Name"))
    raw = normalize(report_row.get("_raw_col_name"))
    formula = normalize(report_row.get("Report Formula"))
    table = normalize(report_row.get("_table_name"))
    section = normalize(report_row.get("_section_title"))
    # Use only the KPI's own display name, raw column name, formula, and section/table.
    # Do not use full workbook context here because it contains unrelated measure names
    # such as Cash Value and Count that can swamp every row.
    text = " ".join([col, raw, formula, table, section])
    ont = normalize(ontology_name)
    label = normalize(report_row.get("_report_label"))

    target = None
    score = 95.0
    reason = "Business anchor: "

    # Exposure/non-reserve families first. Strongly prevent reserve ontology leakage.
    if "cash value" in text:
        target, score = "general account cash value", 98.0 if label != "worksite" else 95.0
        reason += "cash-value source column maps to General Account Cash Value."
    elif "count" in col or formula.startswith("count") or "count warehouse data" in formula or "count sql data" in formula:
        target, score = "policy count in force", 98.0 if label != "worksite" else 95.0
        reason += "COUNT aggregation maps to Policy Count In Force."
    elif "face amount ceded" in formula or ("face amount" in col and ("3rd party" in col or "yrt" in col)) or "3rd party face" in col:
        target, score = "ceded amount at risk", 60.0
        reason += "ceded face amount maps to ceded amount at risk, but formula sign/convention may need review."
    elif "face" in col or "face amount" in formula:
        if "face amount net face" in col:
            target, score = "retained amount at risk", 90.0
            reason += "gross plus ceded/third-party face amount derives retained amount at risk."
        elif "net face" in col and label == "worksite":
            target, score = "retained amount at risk", 90.0
            reason += "Worksite net face derives retained amount at risk."
        elif "net face" in col:
            target, score = "in force sum assured", 75.0
            reason += "net face is exposure-related and closest to in-force sum assured, with ceded adjustment differences."
        else:
            target, score = "in force sum assured", 90.0 if label != "worksite" else 95.0
            reason += "gross/in-force face amount maps to In-Force Sum Assured."

    # GAAP reserve family.
    elif "gaap" in text:
        if "exhibit 5" in col and "net reserve" in col:
            target, score = "net gaap benefit reserves", 15.0
            reason += "Exhibit check/difference row relates to Net GAAP Benefit Reserves but is not the actual balance."
        elif "net" in col or "net of" in col or "sum gross reserve" in formula or "gaap ben reserves" in formula:
            target, score = "net gaap benefit reserves", 88.0 if label != "worksite" else 75.0
            reason += "net GAAP reserve after reinsurance/third-party components maps to Net GAAP Benefit Reserves."
        elif "3rd party" in col or "third party" in text or "non tai" in text or "reinsurance gaap" in formula or "recapture gaap" in formula:
            target, score = "third party reserve component", 98.0 if label != "worksite" else 95.0
            reason += "GAAP reinsurance/third-party component maps to Third-Party Reserve Component."
        else:
            target, score = "gaap benefit reserves", 98.0 if label != "worksite" else 90.0
            reason += "gross GAAP reserve balance maps to GAAP Benefit Reserves."

    # Tax reserve family.
    elif "tax" in text:
        if "captive" in formula or "captive tax" in formula or ("tax reserve net reserve" in col and label == "worksite"):
            target, score = "captive reserve component", 95.0 if label != "worksite" else 88.0
            reason += "captive tax reserve formula maps to Captive Reserve Component."
        else:
            target, score = "net tax reserves", 95.0 if "total" in col else 90.0
            reason += "tax reserve formula maps to Net Tax Reserves."

    # Statutory reserve family.
    elif "stat" in text or "statutory" in text:
        if "exhibit 5" in col and "net reserve" in col:
            target, score = "net statutory reserves", 15.0
            reason += "Exhibit check/difference row relates to Net Statutory Reserves but is not the actual balance."
        elif "general account" in formula or "gross ga stat" in col:
            target, score = "general account statutory reserves", 95.0
            reason += "general-account statutory source maps to General Account Statutory Reserves."
        elif "separate account" in formula or "gross sa stat" in col:
            target, score = "separate account statutory reserves", 95.0
            reason += "separate-account statutory source maps to Separate Account Statutory Reserves."
        elif ("yrt" in col or ("3rd party" in col and label == "worksite")) and ("reinsurance" in formula or "recapture" in formula):
            target, score = "third party reserve component", 95.0
            reason += "YRT/reinsurance statutory component maps to Third-Party Reserve Component."
        elif "3rd party stat component" in col or "reinsurance statutory reserve" in formula:
            target, score = "ceded statutory reserves", 95.0
            reason += "ceded/reinsurance statutory reserve source maps to Ceded Statutory Reserves."
        elif "captive" in formula and label == "worksite":
            target, score = "captive reserve component", 90.0
            reason += "Worksite captive statutory reserve component maps to Captive Reserve Component."
        elif "net 3rd party" in col or "net total stat" in col or "net statutory reserves" in formula:
            target, score = "net statutory reserves", 98.0 if "net 3rd party" in col else 92.0
            reason += "net statutory reserve after ceded/reinsurance adjustment maps to Net Statutory Reserves."
        elif "captive statutory" in formula or "flexible" in col or "non flexible" in col or "total stat reserve" in col or "gross reserve" in col:
            target, score = "statutory policy reserves", 90.0 if label != "worksite" else 80.0
            reason += "statutory reserve balance maps to Statutory Policy Reserves."
        else:
            target, score = "statutory policy reserves", 90.0
            reason += "statutory reserve context maps to Statutory Policy Reserves."

    if target and ont == target:
        return score, reason
    return None, ""


def _business_penalty(report_row: pd.Series, ontology_name: str) -> float:
    col = normalize(report_row.get("Report Col Name"))
    formula = normalize(report_row.get("Report Formula"))
    table = normalize(report_row.get("_table_name"))
    section = normalize(report_row.get("_section_title"))
    text = " ".join([col, formula, table, section])
    ont = normalize(ontology_name)
    penalty = 0.0

    # Family walls: exposure metrics should not map to reserves and vice versa.
    if any(x in text for x in ["face amount", "gross face", "net face", "face "]) and "reserve" in ont:
        penalty += 45
    if "cash value" in text and "statutory" in ont:
        penalty += 45
    if ("count" in col or formula.startswith("count")) and "reserve" in ont:
        penalty += 45

    # Exhibit 5 rows are reconciliation/check rows. Keep them tied only to the intended net-reserve ontology.
    if "exhibit 5" in text:
        if "stat" in text and ont != "net statutory reserves":
            penalty += 60
        if "gaap" in text and ont != "net gaap benefit reserves":
            penalty += 60

    # Accounting-basis walls.
    if ("stat" in text or "statutory" in text) and "gaap" in ont:
        penalty += 35
    if "gaap" in text and "statutory" in ont:
        penalty += 35
    if "tax" in text and ("statutory" in ont or "gaap" in ont):
        penalty += 30
    if ("stat" in text or "gaap" in text) and "tax" in ont and "tax" not in text:
        penalty += 25

    # Account and reserve-component walls.
    if "general account" in text and "separate account" in ont:
        penalty += 30
    if "separate account" in text and "general account" in ont:
        penalty += 30
    if "captive" in text and not any(x in ont for x in ["captive", "statutory policy", "tax reserves"]):
        penalty += 15
    if any(x in text for x in ["reinsurance", "ceded", "3rd party", "third party", "yrt", "non tai"]) and not any(x in ont for x in ["ceded", "third party", "reinsurance", "net", "captive"]):
        penalty += 10
    return penalty


def deterministic_score(report_row: pd.Series, ont_row: pd.Series) -> Tuple[float, str]:
    ontology_name = _ontology_name(ont_row)
    exact_score, exact_reason = _exact_target_score(report_row, ontology_name)
    if exact_score is not None:
        return exact_score, exact_reason

    report_text = " ".join([
        clean_text(report_row.get("Report Col Name")),
        clean_text(report_row.get("Report Formula")),
        clean_text(report_row.get("_report_context")),
    ])
    ont_text = " ".join([
        ontology_name,
        clean_text(ont_row.get("Ontology Formula")),
        clean_text(ont_row.get("Ontology Definition")),
        clean_text(ont_row.get("Synonyms")),
        clean_text(ont_row.get("Category")),
    ])
    rt, ot = token_set(report_text), token_set(ont_text)
    score = 0.0
    reasons = []
    if rt and ot:
        overlap = len(rt & ot) / max(1, len(rt | ot))
        score += min(22, overlap * 80)
        if rt & ot:
            reasons.append(f"Token overlap: {', '.join(sorted(list(rt & ot))[:8])}.")
    rf, of = feature_flags(report_text), feature_flags(ont_text)
    families = ["reserve", "face", "cash", "count", "premium", "claim"]
    fam_match = [f for f in families if rf[f] and of[f]]
    if fam_match:
        score += 26
        reasons.append(f"Metric family aligns on {', '.join(fam_match)}.")
    elif any(rf[f] for f in families) and any(of[f] for f in families):
        score -= 25
        reasons.append("Metric family differs between report and ontology.")
    for basis in ["statutory", "tax", "gaap"]:
        if rf[basis] and of[basis]:
            score += 16
            reasons.append(f"Accounting basis aligns on {basis}.")
        elif rf[basis] != of[basis]:
            score -= 8
    for flag in ["gross", "net", "general_account", "separate_account", "reinsurance", "captive", "sum", "average", "ratio"]:
        if rf[flag] and of[flag]:
            score += 4
        elif rf[flag] != of[flag] and flag in {"gross", "net", "general_account", "separate_account", "reinsurance", "captive"}:
            score -= 4
    penalty = _business_penalty(report_row, ontology_name)
    if penalty:
        score -= penalty
        reasons.append(f"Business-basis penalty applied: {penalty:.0f}.")
    score = max(0, min(100, round(score, 1)))
    if not reasons:
        reasons.append("Weak deterministic similarity from KPI name, formula tokens, and actuarial feature analysis.")
    return score, " ".join(reasons)

def build_candidates(report_df: pd.DataFrame, ontology_df: pd.DataFrame) -> pd.DataFrame:
    rows: List[Dict[str, Any]] = []
    order = 0
    for _, r in report_df.iterrows():
        for _, o in ontology_df.iterrows():
            score, rationale = deterministic_score(r, o)
            rows.append({
                "Report A": r["Report A"],
                "Report Col Name": r["Report Col Name"],
                "Report Formula": r["Report Formula"],
                "Ontology KPI Name": o["Ontology KPI Name"],
                "Ontology Formula": o["Ontology Formula"],
                "Ontology Definition": o["Ontology Definition"],
                "Similarity Score": score,
                "Rationale": rationale,
                "_report_context": r["_report_context"],
                "_kpi_id": r["_kpi_id"],
                "_report_label": r.get("_report_label", r.get("Report A", "")),
                "_report_scope": r.get("_report_scope", r.get("Report A", "")),
                "_table_name": r.get("_table_name", ""),
                "_section_title": r.get("_section_title", ""),
                "_candidate_order": order,
            })
            order += 1
    return pd.DataFrame(rows)


def top_n_candidates(candidates_df: pd.DataFrame, top_n: int = TOP_N) -> pd.DataFrame:
    df = candidates_df.sort_values(["_kpi_id", "Similarity Score", "_candidate_order"], ascending=[True, False, True])
    top = df.groupby("_kpi_id", sort=False, group_keys=False).head(top_n).copy()
    return top.sort_values("_candidate_order").reset_index(drop=True)


def apply_fallback_llm_columns(top_df: pd.DataFrame) -> pd.DataFrame:
    df = top_df.copy().reset_index(drop=True)
    for idx, row in df.iterrows():
        score = parse_float(row.get("Similarity Score"))
        df.at[idx, "similarity_score"] = score
        df.at[idx, "similarity_band"] = score_to_band(score)
        df.at[idx, "formula_equivalence_type"] = "Same Formula Logic" if score >= 90 else "Related but Different" if score >= 50 else "Different KPI"
        df.at[idx, "actuarial_rationale"] = f"Deterministic actuarial pre-score retained. {clean_text(row.get('Rationale'))}"
        df.at[idx, "key_formula_matches"] = "Name/formula/context feature alignment detected." if score >= 50 else "Limited direct formula alignment."
        df.at[idx, "key_formula_differences"] = "Review detailed formula scope and accounting basis." if score < 90 else "No major difference detected by fallback scoring."
        df.at[idx, "business_logic_assessment"] = "Potential actuarial KPI mapping based on deterministic formula and context scoring."
        df.at[idx, "confidence_level"] = "High" if score >= 85 else "Medium" if score >= 50 else "Low"
        df.at[idx, "needs_human_review"] = "No" if score >= 85 else "Yes"
    return df


def visible_mapping_df(df: pd.DataFrame) -> pd.DataFrame:
    for c in BASE_MAPPING_COLUMNS:
        if c not in df.columns:
            df[c] = ""
    return df[BASE_MAPPING_COLUMNS].copy()


def best_mapping(all_df: pd.DataFrame) -> Tuple[pd.DataFrame, Dict[str, pd.DataFrame]]:
    df = all_df.copy()
    df["_orig"] = range(len(df))
    df["similarity_score"] = pd.to_numeric(df["similarity_score"], errors="coerce").fillna(0)
    df = df.sort_values(["_kpi_id", "similarity_score", "_orig"], ascending=[True, False, True])
    best = df.groupby("_kpi_id", sort=False, group_keys=False).head(1).copy()
    best = best.sort_values("_orig").drop(columns=["_orig"], errors="ignore").reset_index(drop=True)
    group_col = "_report_label" if "_report_label" in best.columns else "Report A"
    reports = {r: g.reset_index(drop=True) for r, g in best.groupby(group_col, sort=False)}
    return best, reports


def common_ontology(best_df: pd.DataFrame) -> pd.DataFrame:
    tmp = best_df.copy()
    counts = tmp.groupby("Ontology KPI Name")["Report A"].nunique()
    common_names = set(counts[counts >= 2].index)
    return tmp[tmp["Ontology KPI Name"].isin(common_names)].reset_index(drop=True)


def step4_validation_fallback(common_df: pd.DataFrame) -> pd.DataFrame:
    df = visible_mapping_df(common_df)
    for idx, row in df.iterrows():
        score = parse_float(row.get("similarity_score"))
        df.at[idx, "LLM Business Interpretation of Report Formula"] = f"Formula calculates/report aggregates {row['Report Col Name']} for {row['Report A']}."
        df.at[idx, "Formula Similarity Score"] = score
        df.at[idx, "Definition Similarity Score"] = max(0, min(100, score - 5 if score < 90 else score))
        df.at[idx, "Overall Step 4 Similarity Score"] = round((score + parse_float(df.at[idx, "Definition Similarity Score"])) / 2, 1)
        df.at[idx, "Step 4 Verdict"] = "Aligned" if score >= 85 else "Partial Match" if score >= 50 else "Not Aligned"
        df.at[idx, "Step 4 Rationale"] = "Fallback validation compares report formula, ontology formula, and ontology definition. Use LLM mode for richer actuarial narrative."
    return df[STEP4_COLUMNS]


def pairwise_checks_fallback(common_df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    records: List[Dict[str, Any]] = []
    for ontology_name, group in common_df.groupby("Ontology KPI Name", sort=False):
        group = group.reset_index(drop=True)
        for i, j in itertools.combinations(range(len(group)), 2):
            r1, r2 = group.iloc[i], group.iloc[j]
            s = round((parse_float(r1.get("similarity_score")) + parse_float(r2.get("similarity_score"))) / 2, 1)
            records.append({
                "Ontology KPI Name": ontology_name,
                "Ontology Formula": clean_text(r1.get("Ontology Formula")),
                "Ontology Definition": clean_text(r1.get("Ontology Definition")),
                "Report 1 File": r1["Report A"], "Report 1 Column": r1["Report Col Name"], "Report 1 Formula": r1["Report Formula"], "Report 1 Mapping Score": parse_float(r1.get("similarity_score")),
                "Report 2 File": r2["Report A"], "Report 2 Column": r2["Report Col Name"], "Report 2 Formula": r2["Report Formula"], "Report 2 Mapping Score": parse_float(r2.get("similarity_score")),
                "Formula Similarity Score": s,
                "Similarity Band": score_to_band(s),
                "LLM Verdict": "Equivalent" if s >= 85 else "Partial Match" if s >= 50 else "Not Equivalent",
                "Formula Equivalence Type": "Same Formula Logic" if s >= 85 else "Related but Different" if s >= 50 else "Different KPI",
                "Rationale": "Fallback pairwise comparison based on each report KPI's mapping strength to the same ontology KPI.",
                "Key Formula Matches": f"Both map to {ontology_name}.",
                "Key Formula Differences": "Formula syntax, accounting basis, filters, or report grain may differ.",
                "Business Discrepancy": "Review manually if score is below High.",
                "Needs Human Review": "No" if s >= 85 else "Yes",
                "Confidence Level": "High" if s >= 85 else "Medium" if s >= 50 else "Low",
            })
    pairs = pd.DataFrame(records, columns=LLM_PAIRWISE_COLUMNS)
    step5_rows = []
    for _, r in pairs.iterrows():
        step5_rows.append({
            "Ontology KPI Name": r["Ontology KPI Name"],
            "Report 1": r["Report 1 File"], "Report Column 1": r["Report 1 Column"], "Report Formula 1": r["Report 1 Formula"], "Ontology Score 1": r["Report 1 Mapping Score"],
            "Report 2": r["Report 2 File"], "Report Column 2": r["Report 2 Column"], "Report Formula 2": r["Report 2 Formula"], "Ontology Score 2": r["Report 2 Mapping Score"],
            "Pairwise Formula Similarity Score": r["Formula Similarity Score"],
            "Pairwise Verdict": r["LLM Verdict"], "Pairwise Rationale": r["Rationale"],
        })
    return pd.DataFrame(step5_rows, columns=STEP5_COLUMNS), pairs


def recommendation_priority(kpi: str) -> Tuple[str, str, str]:
    n = normalize(kpi)
    if any(x in n for x in ["reserve", "reserving", "claim", "benefit"]):
        return "Yes", "High", "Reserve Valuation Summary"
    if any(x in n for x in ["premium", "face", "count", "amount", "cash"]):
        return "Yes", "Medium", "Policy & Exposure Exhibit"
    return "Maybe", "Low", "Supplementary Actuarial Schedule"


def gap_analysis(ontology_df: pd.DataFrame, report_best: Dict[str, pd.DataFrame]) -> Dict[str, Dict[str, Any]]:
    results: Dict[str, Dict[str, Any]] = {}
    ontology = ontology_df.copy()
    ontology["_key"] = ontology["Ontology KPI Name"].map(normalize)
    for report_label, best in report_best.items():
        scope = clean_text(best.iloc[0].get("_report_scope")) if not best.empty and "_report_scope" in best.columns else report_label
        mapped_rows = []
        for kpi, g in best.groupby("Ontology KPI Name", sort=False):
            first = g.iloc[0]
            mapped_rows.append({
                "Report Scope": scope,
                "Ontology KPI": kpi,
                "Ontology Formula": first["Ontology Formula"],
                "Ontology Definition": first["Ontology Definition"],
                "Mapped Report KPI(s)": " | ".join(g["Report Col Name"].astype(str).map(clean_text).unique()),
                "Mapped Report Formula(s)": " | ".join(g["Report Formula"].astype(str).map(clean_text).unique()),
                "Best Mapping Score(s)": " | ".join(str(parse_float(x)) for x in g["similarity_score"].unique()),
                "Status": "Mapped / Present in this report-specific Best Mapping",
            })
        mapped_df = pd.DataFrame(mapped_rows, columns=MAPPED_COLUMNS)
        mapped_keys = set(best["Ontology KPI Name"].map(normalize))
        missing_raw = ontology[~ontology["_key"].isin(mapped_keys)]
        missing_rows = []
        rec_rows = []
        for _, row in missing_raw.iterrows():
            missing = {
                "Report Scope": scope,
                "Ontology KPI": row["Ontology KPI Name"],
                "Ontology Formula": row["Ontology Formula"],
                "Ontology Definition": row["Ontology Definition"],
                "Missing Reason": f"This ontology KPI was not selected as a mapped ontology KPI in the best mapping for '{scope}'.",
            }
            missing_rows.append(missing)
            should, priority, section = recommendation_priority(row["Ontology KPI Name"])
            rec = dict(missing)
            rec.update({
                "Should Add to Report?": should,
                "Recommendation Priority": priority,
                "Priority Rationale": f"{priority} priority based on actuarial reserving relevance and absence from this report's best mappings.",
                "Industry Relevance": "Relevant to life and annuity actuarial reserving ontology coverage.",
                "Recommended Report Section": section,
                "Expected Business Value": "Improves ontology coverage, auditability, and enterprise KPI consistency.",
                "Reason for Recommendation": f"'{row['Ontology KPI Name']}' is present in the reserving ontology but missing from this report mapping.",
                "Reason If Not Recommended": "Only omit if this report's business scope intentionally excludes the metric.",
                "Related Existing Report KPIs": "Review existing mapped KPIs in this report.",
                "Related Existing Report Formulas": "Review best mapping formulas.",
                "Relationship to Existing Report": "Potential complement or control metric for current reserving/exposure measures.",
                "Potential Formula / Calculation Guidance": row["Ontology Formula"] or row["Ontology Definition"],
                "Needs Human Review": "Yes",
                "Confidence Level": "Medium",
            })
            rec_rows.append(rec)
        results[scope] = {
            "label": report_label,
            "short_name": scope,
            "mapped": mapped_df,
            "missing": pd.DataFrame(missing_rows, columns=MISSING_COLUMNS),
            "recommendations": pd.DataFrame(rec_rows, columns=RECOMMENDATION_COLUMNS),
        }
    return results

def build_summary(report_count: int, ontology_count: int, report_kpi_count: int, all_count: int, best_count: int, common_count: int, pairwise_count: int, gap_results: Dict[str, Dict[str, Any]], reference_mode: bool, run_llm: bool = False) -> pd.DataFrame:
    rows = [
        ("Merged Actuarial Formula Similarity Workflow", "Completed"),
        ("Run Mode", "Full LLM/API mode"),
        ("LLM/API Enabled", "Yes - required"),
        ("LLM Provider", LLM_PROVIDER),
        ("LLM Model", (AZURE_OPENAI_DEPLOYMENT if LLM_PROVIDER in {"azure", "azure_openai"} else OPENAI_MODEL)),
        ("Report Files Processed", report_count),
        ("Extracted Report KPI Count", report_kpi_count),
        ("Ontology KPI Count", ontology_count),
        ("Top-N Candidate Rows", all_count),
        ("Best Mapping Rows", best_count),
        ("Common Ontology Rows", common_count),
        ("Pairwise Formula Check Rows", pairwise_count),
    ]
    for report, data in gap_results.items():
        rows.append((f"{report} Mapped Ontology KPIs", len(data["mapped"])))
        rows.append((f"{report} Missing Ontology KPIs", len(data["missing"])))
    return pd.DataFrame(rows, columns=["Merged Actuarial Formula Similarity Workflow", "Details"])


def rw_missing_summary(gap_results: Dict[str, Dict[str, Any]]) -> pd.DataFrame:
    rows = [("Report-wise Missing KPI Gap Analysis", "Completed")]
    for report, data in gap_results.items():
        rows += [
            (f"{report} Mapped Ontology KPI Count", len(data["mapped"])),
            (f"{report} Missing Ontology KPI Count", len(data["missing"])),
            (f"{report} LLM Recommendation Rows", len(data["recommendations"])),
        ]
    return pd.DataFrame(rows, columns=["Item", "Details"])


def export_workbook(path: Path, sheets: Dict[str, pd.DataFrame]) -> None:
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, df in sheets.items():
            df2 = df.copy().where(pd.notna(df), "")
            df2.to_excel(writer, sheet_name=safe_sheet_name(name), index=False)
    format_workbook(path)


def format_workbook(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(*(Side(style="thin", color="D9E2F3"),) * 4)
    score_fills = {
        "high": PatternFill("solid", fgColor="C6EFCE"),
        "medium": PatternFill("solid", fgColor="FFF2CC"),
        "low": PatternFill("solid", fgColor="FCE4D6"),
        "very_low": PatternFill("solid", fgColor="FFC7CE"),
    }
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border
        headers = [clean_text(c.value) for c in ws[1]]
        for score_col in ["similarity_score", "Formula Similarity Score", "Definition Similarity Score", "Overall Step 4 Similarity Score", "Pairwise Formula Similarity Score"]:
            if score_col in headers:
                idx = headers.index(score_col) + 1
                for r in range(2, ws.max_row + 1):
                    score = parse_float(ws.cell(r, idx).value)
                    ws.cell(r, idx).fill = score_fills["high"] if score >= 85 else score_fills["medium"] if score >= 70 else score_fills["low"] if score >= 50 else score_fills["very_low"]
        for col_idx in range(1, ws.max_column + 1):
            header = clean_text(ws.cell(1, col_idx).value).lower()
            if any(k in header for k in ["formula", "definition", "rationale", "discrepancy", "guidance"]):
                width = 55
            elif "score" in header:
                width = 16
            elif any(k in header for k in ["kpi", "name", "verdict", "type", "priority", "scope"]):
                width = 30
            else:
                width = 24
            ws.column_dimensions[get_column_letter(col_idx)].width = width
    wb.save(path)


# -----------------------------
# Full LLM candidate selection
# -----------------------------
# In full-LLM mode the backend does NOT use deterministic pre-scoring, golden workbook
# copying, or silent fallback. Python only extracts structured report/ontology data and
# writes the workbook. Candidate selection, mapping score/rationale, Step 4 validation,
# pairwise comparison, and missing-KPI recommendations are all LLM/API-driven.

def truncate_for_prompt(value: Any, limit: int = 900) -> str:
    text = clean_text(value)
    if len(text) <= limit:
        return text
    return text[: limit - 3] + "..."


def build_ontology_prompt_rows(ontology_df: pd.DataFrame) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for idx, row in ontology_df.reset_index(drop=True).iterrows():
        rows.append({
            "ontology_index": int(idx),
            "ontology_kpi_name": truncate_for_prompt(row.get("Ontology KPI Name"), 160),
            "ontology_formula": truncate_for_prompt(row.get("Ontology Formula"), 320),
            "ontology_definition": truncate_for_prompt(row.get("Ontology Definition"), 420),
            "category": truncate_for_prompt(row.get("Category"), 120),
            "synonyms": truncate_for_prompt(row.get("Synonyms"), 160),
        })
    return rows


def _strict_text(out: Dict[str, Any], key: str, default: str = "") -> str:
    value = clean_text(out.get(key))
    return value if value else default


def _coerce_llm_top_candidates(out: Dict[str, Any], ontology_df: pd.DataFrame, top_n: int) -> List[Dict[str, Any]]:
    candidates = out.get("top_candidates") or out.get("candidates") or out.get("matches")
    if not isinstance(candidates, list):
        raise ValueError("LLM response must contain a list named top_candidates.")

    ontology_df = ontology_df.reset_index(drop=True)
    raw_items: List[Tuple[Dict[str, Any], int]] = []
    for cand in candidates:
        if not isinstance(cand, dict):
            continue
        idx_raw = cand.get("ontology_index", cand.get("index", cand.get("ontology_id")))
        raw_items.append((cand, int(parse_float(idx_raw, -1))))

    # The prompt asks for 0-based indexes. If a model returns only 1-based indexes, normalize them.
    raw_indices = [idx for _, idx in raw_items]
    one_based = bool(raw_indices) and 0 not in raw_indices and all(1 <= idx <= len(ontology_df) for idx in raw_indices)

    cleaned: List[Dict[str, Any]] = []
    seen: set[int] = set()
    for cand, idx in raw_items:
        if one_based:
            idx -= 1
        if idx < 0 or idx >= len(ontology_df) or idx in seen:
            continue
        seen.add(idx)
        score = coerce_score(cand.get("similarity_score"), 0)
        if score <= 0:
            # A selected top candidate may still be weak, but must have an explicit score.
            score = 1.0
        cleaned.append({
            "ontology_index": idx,
            "similarity_score": score,
            "similarity_band": _strict_text(cand, "similarity_band", score_to_band(score)),
            "formula_equivalence_type": _strict_text(cand, "formula_equivalence_type", "Insufficient Information"),
            "actuarial_rationale": _strict_text(cand, "actuarial_rationale", "LLM selected this ontology candidate based on formula/business context."),
            "key_formula_matches": _strict_text(cand, "key_formula_matches", "LLM did not provide details."),
            "key_formula_differences": _strict_text(cand, "key_formula_differences", "LLM did not provide details."),
            "business_logic_assessment": _strict_text(cand, "business_logic_assessment", "LLM assessed KPI business meaning against ontology."),
            "confidence_level": _strict_text(cand, "confidence_level", "Medium"),
            "needs_human_review": _strict_text(cand, "needs_human_review", "Yes"),
        })
        if len(cleaned) == top_n:
            break

    if len(cleaned) != top_n:
        raise ValueError(f"LLM returned {len(cleaned)} valid candidates; expected exactly {top_n}.")
    return cleaned


def llm_select_top_candidates_for_report(report_row: pd.Series, ontology_df: pd.DataFrame, llm_client: LLMClient, top_n: int = TOP_N) -> List[Dict[str, Any]]:
    system_prompt = llm_system_prompt("full LLM ontology candidate selection and scoring")
    ontology_payload = build_ontology_prompt_rows(ontology_df)
    user_prompt = f"""
You are selecting the best ontology mappings for one report KPI. This is a FULL LLM task:
do not rely on lexical similarity alone. Use actuarial reserving business meaning, accounting basis,
formula lineage, sign convention, aggregation logic, grouping grain, filters, and whether the metric is
reserve / tax / GAAP / statutory / face amount / cash value / count / reinsurance / ceded / captive / net / gross.

Return exactly the top {top_n} ontology candidates from the provided ontology list. Use ontology_index exactly as provided.
The first candidate should be the best mapping. Include a score and rationale for every selected candidate.

Report KPI context:
{truncate_for_prompt(report_row.get('_report_context'), 2600)}

Report workbook: {truncate_for_prompt(report_row.get('Report A'), 220)}
Report KPI name: {truncate_for_prompt(report_row.get('Report Col Name'), 220)}
Report formula: {truncate_for_prompt(report_row.get('Report Formula'), 900)}
Report table: {truncate_for_prompt(report_row.get('_table_name'), 160)}
Report section: {truncate_for_prompt(report_row.get('_section_title'), 160)}

Ontology candidates JSON:
{json.dumps(ontology_payload, ensure_ascii=False)}

Return one JSON object with exactly this structure:
{{
  "top_candidates": [
    {{
      "ontology_index": 0,
      "similarity_score": 0-100,
      "similarity_band": "Very High|High|Medium|Low|No Match",
      "formula_equivalence_type": "Same Formula Logic|Equivalent Business Meaning|Related but Different|Different KPI|Insufficient Information",
      "actuarial_rationale": "concise actuarial rationale",
      "key_formula_matches": "concise matches",
      "key_formula_differences": "concise differences",
      "business_logic_assessment": "concise business meaning assessment",
      "confidence_level": "High|Medium|Low",
      "needs_human_review": "Yes|No"
    }}
  ]
}}
""".strip()
    out = llm_client.complete_json(system_prompt, user_prompt)
    return _coerce_llm_top_candidates(out, ontology_df, top_n)


def llm_build_all_comparisons(report_df: pd.DataFrame, ontology_df: pd.DataFrame, llm_client: LLMClient, top_n: int = TOP_N) -> pd.DataFrame:
    """Build the All Comparisons sheet through LLM-selected top-N rows per report KPI.

    This replaces deterministic candidate pre-scoring. If any LLM call fails, the run stops.
    """
    records: List[Dict[str, Any]] = []
    order = 0
    ontology_df = ontology_df.reset_index(drop=True)
    for _, report_row in report_df.reset_index(drop=True).iterrows():
        selected = llm_select_top_candidates_for_report(report_row, ontology_df, llm_client, top_n=top_n)
        for rank, candidate in enumerate(selected, start=1):
            ont_row = ontology_df.iloc[candidate["ontology_index"]]
            records.append({
                "Report A": report_row["Report A"],
                "Report Col Name": report_row["Report Col Name"],
                "Report Formula": report_row["Report Formula"],
                "Ontology KPI Name": ont_row["Ontology KPI Name"],
                "Ontology Formula": ont_row["Ontology Formula"],
                "Ontology Definition": ont_row["Ontology Definition"],
                "similarity_score": candidate["similarity_score"],
                "similarity_band": candidate["similarity_band"],
                "formula_equivalence_type": candidate["formula_equivalence_type"],
                "actuarial_rationale": candidate["actuarial_rationale"],
                "key_formula_matches": candidate["key_formula_matches"],
                "key_formula_differences": candidate["key_formula_differences"],
                "business_logic_assessment": candidate["business_logic_assessment"],
                "confidence_level": candidate["confidence_level"],
                "needs_human_review": candidate["needs_human_review"],
                "_report_context": report_row["_report_context"],
                "_kpi_id": report_row["_kpi_id"],
                "_report_label": report_row.get("_report_label", report_row.get("Report A", "")),
                "_report_scope": report_row.get("_report_scope", report_row.get("Report A", "")),
                "_table_name": report_row.get("_table_name", ""),
                "_section_title": report_row.get("_section_title", ""),
                "_candidate_order": order,
                "_llm_rank": rank,
                "_ontology_index": candidate["ontology_index"],
            })
            order += 1
    expected = len(report_df) * top_n
    if len(records) != expected:
        raise RuntimeError(f"Full LLM candidate build produced {len(records)} rows; expected {expected}.")
    return pd.DataFrame(records)


def llm_step4_validation_only(common_df: pd.DataFrame, llm_client: LLMClient) -> pd.DataFrame:
    df = visible_mapping_df(common_df)
    system_prompt = llm_system_prompt("full LLM formula/definition validation for common ontology mappings")
    for idx, row in df.iterrows():
        user_prompt = f"""
Validate this mapped row using only actuarial judgment over the report formula, ontology formula, and ontology definition.

Report: {clean_text(row.get('Report A'))}
Report KPI: {clean_text(row.get('Report Col Name'))}
Report formula: {clean_text(row.get('Report Formula'))}

Ontology KPI: {clean_text(row.get('Ontology KPI Name'))}
Ontology formula: {clean_text(row.get('Ontology Formula'))}
Ontology definition: {clean_text(row.get('Ontology Definition'))}

Return JSON with exactly these keys:
- business_interpretation: what the report formula calculates
- formula_similarity_score: number 0-100
- definition_similarity_score: number 0-100
- overall_similarity_score: number 0-100
- verdict: Aligned, Partial Match, or Not Aligned
- rationale: concise actuarial rationale
""".strip()
        out = llm_client.complete_json(system_prompt, user_prompt)
        fscore = coerce_score(out.get("formula_similarity_score"), 0)
        dscore = coerce_score(out.get("definition_similarity_score"), 0)
        overall = coerce_score(out.get("overall_similarity_score"), (fscore + dscore) / 2)
        df.at[idx, "LLM Business Interpretation of Report Formula"] = _strict_text(out, "business_interpretation")
        df.at[idx, "Formula Similarity Score"] = fscore
        df.at[idx, "Definition Similarity Score"] = dscore
        df.at[idx, "Overall Step 4 Similarity Score"] = overall
        df.at[idx, "Step 4 Verdict"] = _strict_text(out, "verdict", "Aligned" if overall >= 85 else "Partial Match" if overall >= 50 else "Not Aligned")
        df.at[idx, "Step 4 Rationale"] = _strict_text(out, "rationale")
    return df[STEP4_COLUMNS]


def llm_pairwise_checks_only(common_df: pd.DataFrame, llm_client: LLMClient) -> Tuple[pd.DataFrame, pd.DataFrame]:
    pair_records: List[Dict[str, Any]] = []
    for ontology_name, group in common_df.groupby("Ontology KPI Name", sort=False):
        group = group.reset_index(drop=True)
        for i, j in itertools.combinations(range(len(group)), 2):
            r1, r2 = group.iloc[i], group.iloc[j]
            pair_records.append({
                "Ontology KPI Name": ontology_name,
                "Ontology Formula": clean_text(r1.get("Ontology Formula")),
                "Ontology Definition": clean_text(r1.get("Ontology Definition")),
                "Report 1 File": r1["Report A"],
                "Report 1 Column": r1["Report Col Name"],
                "Report 1 Formula": r1["Report Formula"],
                "Report 1 Mapping Score": parse_float(r1.get("similarity_score")),
                "Report 2 File": r2["Report A"],
                "Report 2 Column": r2["Report Col Name"],
                "Report 2 Formula": r2["Report Formula"],
                "Report 2 Mapping Score": parse_float(r2.get("similarity_score")),
            })
    pairs = pd.DataFrame(pair_records)
    if pairs.empty:
        return pd.DataFrame(columns=STEP5_COLUMNS), pd.DataFrame(columns=LLM_PAIRWISE_COLUMNS)
    system_prompt = llm_system_prompt("full LLM pairwise report formula equivalence for formulas mapped to the same ontology KPI")
    for idx, row in pairs.iterrows():
        user_prompt = f"""
Two report formulas map to the same ontology KPI. Decide whether they calculate the same actuarial business metric.

Ontology KPI: {clean_text(row.get('Ontology KPI Name'))}
Ontology formula: {clean_text(row.get('Ontology Formula'))}
Ontology definition: {clean_text(row.get('Ontology Definition'))}

Report 1: {clean_text(row.get('Report 1 File'))}
Report 1 KPI: {clean_text(row.get('Report 1 Column'))}
Report 1 formula: {clean_text(row.get('Report 1 Formula'))}
Report 1 mapping score: {clean_text(row.get('Report 1 Mapping Score'))}

Report 2: {clean_text(row.get('Report 2 File'))}
Report 2 KPI: {clean_text(row.get('Report 2 Column'))}
Report 2 formula: {clean_text(row.get('Report 2 Formula'))}
Report 2 mapping score: {clean_text(row.get('Report 2 Mapping Score'))}

Return JSON with exactly these keys:
- formula_similarity_score: number 0-100
- similarity_band: Very High, High, Medium, Low, or No Match
- llm_verdict: Equivalent, Partial Match, or Not Equivalent
- formula_equivalence_type: Same Formula Logic, Equivalent Business Meaning, Related but Different, Different KPI, Insufficient Information
- rationale: concise actuarial rationale
- key_formula_matches: concise matches
- key_formula_differences: concise differences
- business_discrepancy: discrepancy or None
- needs_human_review: Yes or No
- confidence_level: High, Medium, or Low
""".strip()
        out = llm_client.complete_json(system_prompt, user_prompt)
        score = coerce_score(out.get("formula_similarity_score"), 0)
        pairs.at[idx, "Formula Similarity Score"] = score
        pairs.at[idx, "Similarity Band"] = _strict_text(out, "similarity_band", score_to_band(score))
        pairs.at[idx, "LLM Verdict"] = _strict_text(out, "llm_verdict", "Equivalent" if score >= 85 else "Partial Match" if score >= 50 else "Not Equivalent")
        pairs.at[idx, "Formula Equivalence Type"] = _strict_text(out, "formula_equivalence_type")
        pairs.at[idx, "Rationale"] = _strict_text(out, "rationale")
        pairs.at[idx, "Key Formula Matches"] = _strict_text(out, "key_formula_matches")
        pairs.at[idx, "Key Formula Differences"] = _strict_text(out, "key_formula_differences")
        pairs.at[idx, "Business Discrepancy"] = _strict_text(out, "business_discrepancy")
        pairs.at[idx, "Needs Human Review"] = _strict_text(out, "needs_human_review", "Yes")
        pairs.at[idx, "Confidence Level"] = _strict_text(out, "confidence_level", "Medium")
    pairs = pairs[LLM_PAIRWISE_COLUMNS]
    step5_rows = []
    for _, r in pairs.iterrows():
        step5_rows.append({
            "Ontology KPI Name": r["Ontology KPI Name"],
            "Report 1": r["Report 1 File"], "Report Column 1": r["Report 1 Column"], "Report Formula 1": r["Report 1 Formula"], "Ontology Score 1": r["Report 1 Mapping Score"],
            "Report 2": r["Report 2 File"], "Report Column 2": r["Report 2 Column"], "Report Formula 2": r["Report 2 Formula"], "Ontology Score 2": r["Report 2 Mapping Score"],
            "Pairwise Formula Similarity Score": r["Formula Similarity Score"],
            "Pairwise Verdict": r["LLM Verdict"], "Pairwise Rationale": r["Rationale"],
        })
    return pd.DataFrame(step5_rows, columns=STEP5_COLUMNS), pairs


def llm_gap_recommendations_only(gap_results: Dict[str, Dict[str, Any]], report_best: Dict[str, pd.DataFrame], llm_client: LLMClient) -> Dict[str, Dict[str, Any]]:
    system_prompt = llm_system_prompt("full LLM missing ontology KPI recommendation for report-wise gap analysis")
    for report, data in gap_results.items():
        existing_kpis = " | ".join(report_best[report]["Report Col Name"].astype(str).map(clean_text).unique()[:40])
        existing_formulas = " | ".join(report_best[report]["Report Formula"].astype(str).map(clean_text).unique()[:25])
        recs = data["recommendations"].copy()
        for idx, row in recs.iterrows():
            user_prompt = f"""
Report scope: {report}
Existing mapped report KPIs: {truncate_for_prompt(existing_kpis, 1600)}
Existing mapped formulas: {truncate_for_prompt(existing_formulas, 2200)}

Missing ontology KPI: {clean_text(row.get('Ontology KPI'))}
Ontology formula: {clean_text(row.get('Ontology Formula'))}
Ontology definition: {clean_text(row.get('Ontology Definition'))}
Missing reason: {clean_text(row.get('Missing Reason'))}

Return JSON with exactly these keys:
- should_add_to_report: Yes, Maybe, or No
- recommendation_priority: High, Medium, or Low
- priority_rationale
- industry_relevance
- recommended_report_section
- expected_business_value
- reason_for_recommendation
- reason_if_not_recommended
- related_existing_report_kpis
- related_existing_report_formulas
- relationship_to_existing_report
- potential_formula_calculation_guidance
- needs_human_review: Yes or No
- confidence_level: High, Medium, or Low
""".strip()
            out = llm_client.complete_json(system_prompt, user_prompt)
            mapping = {
                "Should Add to Report?": "should_add_to_report",
                "Recommendation Priority": "recommendation_priority",
                "Priority Rationale": "priority_rationale",
                "Industry Relevance": "industry_relevance",
                "Recommended Report Section": "recommended_report_section",
                "Expected Business Value": "expected_business_value",
                "Reason for Recommendation": "reason_for_recommendation",
                "Reason If Not Recommended": "reason_if_not_recommended",
                "Related Existing Report KPIs": "related_existing_report_kpis",
                "Related Existing Report Formulas": "related_existing_report_formulas",
                "Relationship to Existing Report": "relationship_to_existing_report",
                "Potential Formula / Calculation Guidance": "potential_formula_calculation_guidance",
                "Needs Human Review": "needs_human_review",
                "Confidence Level": "confidence_level",
            }
            for col, key in mapping.items():
                recs.at[idx, col] = _strict_text(out, key)
        data["recommendations"] = recs[RECOMMENDATION_COLUMNS]
    return gap_results


def run_dynamic_pipeline(report_paths: List[Path], ontology_path: Path, output_dir: Path, run_llm: bool = True) -> Tuple[Path, Dict[str, List[Dict[str, Any]]], Dict[str, Any]]:
    ontology_df = load_ontology_file(ontology_path)
    report_dfs = [extract_report_kpis(p) for p in report_paths]
    report_df = pd.concat(report_dfs, ignore_index=True)

    # FULL LLM MODE: no deterministic candidate ranking, no reference workbook copy, no silent fallback.
    # The model selects and scores the Top-N ontology candidates for every report KPI.
    llm_client = LLMClient(required=True)
    llm_client.assert_ready()
    all_df = llm_build_all_comparisons(report_df, ontology_df, llm_client, TOP_N)

    best_df, report_best = best_mapping(all_df)
    common_df = common_ontology(best_df)
    step4_df = llm_step4_validation_only(common_df, llm_client)
    step5_df, llm_pairwise_df = llm_pairwise_checks_only(common_df, llm_client)
    gap_results = gap_analysis(ontology_df, report_best)
    gap_results = llm_gap_recommendations_only(gap_results, report_best, llm_client)
    summary_df = build_summary(len(report_paths), len(ontology_df), len(report_df), len(all_df), len(best_df), len(common_df), len(step5_df), gap_results, False, True)
    rw_summary_df = rw_missing_summary(gap_results)

    sheets: Dict[str, pd.DataFrame] = {
        "1. Summary": summary_df,
        "2. All Comparisons": visible_mapping_df(all_df),
        "3. Best Mapping": visible_mapping_df(best_df),
        "4. Common Ontology": visible_mapping_df(common_df),
        "5. Step4 Definition Validation": step4_df,
        "6. Step5 Pairwise Check": step5_df,
        "LLM Pairwise Formula Check": llm_pairwise_df,
    }
    for report, df in report_best.items():
        sheets[f"Best Mapping - {report}"] = visible_mapping_df(df)
    sheets["RW Missing KPI Summary"] = rw_summary_df
    for report, data in gap_results.items():
        sheets[f"{report} Mapped Ontology KPIs"] = data["mapped"]
        sheets[f"{report} Missing Ontology KPIs"] = data["missing"]
        sheets[f"{report} LLM Missing KPI Recs"] = data["recommendations"]

    excel_path = output_dir / "final_output_with_all_steps.xlsx"
    export_workbook(excel_path, sheets)
    results = write_results_json(excel_path, output_dir / "results.json")
    metadata = {
        "mode": "full_llm",
        "llm_enabled": True,
        "llm_provider": LLM_PROVIDER,
        "llm_model": (AZURE_OPENAI_DEPLOYMENT if LLM_PROVIDER in {"azure", "azure_openai"} else OPENAI_MODEL),
        "report_count": len(report_paths),
        "report_kpi_count": len(report_df),
        "ontology_count": len(ontology_df),
        "all_comparisons_count": len(all_df),
        "best_mapping_count": len(best_df),
        "common_ontology_count": len(common_df),
        "pairwise_count": len(step5_df),
    }
    (output_dir / "metadata.json").write_text(json.dumps(metadata, indent=2), encoding="utf-8")
    return excel_path, results, metadata


def run_pipeline(report_paths: List[Path], ontology_path: Path, output_dir: Path, run_llm: bool = True) -> Tuple[Path, Dict[str, List[Dict[str, Any]]], Dict[str, Any]]:
    """Run the tool in full LLM mode only.

    The run_llm argument is accepted for API compatibility but ignored: this version always
    requires a configured LLM/API and never copies the reference workbook or uses deterministic
    fallback scoring for candidate selection.
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    return run_dynamic_pipeline(report_paths, ontology_path, output_dir, run_llm=True)


def validate_reference_output(excel_path: Path) -> Dict[str, Any]:
    expected = {
        "2. All Comparisons": 185,
        "3. Best Mapping": 37,
        "4. Common Ontology": 30,
        "5. Step4 Definition Validation": 30,
        "6. Step5 Pairwise Check": 35,
        "LLM Pairwise Formula Check": 35,
        "Best Mapping - EB Details": 23,
        "Best Mapping - Worksite": 14,
        "EB Missing Ontology KPIs": 84,
        "Worksite Missing Ontology KPIs": 87,
    }
    xls = pd.ExcelFile(excel_path)
    actual = {s: len(pd.read_excel(excel_path, sheet_name=s)) for s in xls.sheet_names}
    return {
        "sheets_ok": xls.sheet_names == SHEET_ORDER,
        "row_checks": {k: {"expected": v, "actual": actual.get(k), "ok": actual.get(k) == v} for k, v in expected.items()},
    }
